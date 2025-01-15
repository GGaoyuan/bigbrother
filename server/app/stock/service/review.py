import numpy as np
from concurrent.futures import ThreadPoolExecutor
import app.stock.dao.akshare_wrapper as akshare
import app.stock.dao.adata_wrapper as adata
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from pandas import DataFrame
from ..dao.field_mapper import *
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed


def get_core_board():
    today = pd.Timestamp.now().strftime("%Y%m%d")
    # 涨停
    limit_up_df = akshare.get_stock_limit_up(today)
    # 跌停
    limit_down_df = akshare.get_stock_limit_down(today)
    # stock_list = limit_up_df[STOCK_CODE].tolist()
    # print(stock_list)
    # with ThreadPoolExecutor(max_workers=len(limit_up_df)) as executor:
    #     futures = {executor.submit(adata.get_stock_industry, row[STOCK_CODE]): row for _, row in limit_up_df.iterrows()}
    #     for future in as_completed(futures):
    #         try:
    #             result = future.result()
    #             result[STOCK_NAME] = futures[future][STOCK_NAME]
    #         except Exception as e:
    #             print(f"线程执行错误: {e}")

    results = []
    max_workers = len(limit_up_df)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {}
        for index, row in limit_up_df.iterrows():
            stock_code = row[STOCK_CODE]
            futures[executor.submit(adata.get_stock_core_concept, stock_code)] = row
            futures[executor.submit(adata.get_stock_industry, stock_code)] = row
        # 等待所有任务完成并处理结果
        for future in as_completed(futures):
            try:
                df = future.result()
                # df[STOCK_NAME] = futures[future][STOCK_NAME]
                results.append(df)
            except Exception as e:
                print(f"Error occurred: {e}")
    print("limit_up_df")



    # industry_boards = get_industry_board_df()
    # if industry_boards.empty:
    #     print(f"获取行业板块失败")
    #     return pd.DataFrame()
    # # 获取板块名称列表
    # industry_names = industry_boards["板块名称"].tolist()


    # df1 = adata.get_stock_concept('300033')



    # print("")
    # df2 = adata.get_stock_concept2('300033')
    # print("")
    # df3 = adata.get_stock_industry('300033')
    # print("")
    # df4 = adata.get_stock_district('300033')
    # print("")

    """
    today = pd.Timestamp.now().strftime("%Y%m%d")
    # 涨停
    limit_up = wrapper.get_limit_up_df(today)
    # 跌停
    limit_down = wrapper.get_limit_down_df(today)
    # 行业板块
    industry_df = wrapper.get_industry_board_components_detail_df()
    # 概念板块
    concept_df = wrapper.get_concept_board_components_detail_df()
    board_df = pd.merge(industry_df, concept_df, on='代码', how='outer')
    limit_up_stock = []
    for index, row in limit_up.iterrows():
        df = board_df[board_df['代码'] == row['代码']]
        limit_up_stock.append(df)
    df = pd.concat(limit_up_stock, ignore_index=True)
    print("")
"""


    # if not concept_board_detail_list:
    #     print("失败4")
    #     return














class ReviewService:
    def main(self):
        self.daily_review()


    def pre_market_review(self):
        pass

    def lunch_break_review(self):
        """
        午间休息
        """

        pass

    def after_market_review(self):
        """
        盘后休息
        """
        pass


    def daily_review(self):
        pass
        """
        data_str = '20241226'
        def get_limit_up_data():
            return akshare.get_limit_up_data(data_str)
        def get_limit_down_data():
            return akshare.get_limit_down_data(data_str)
        def get_industry_sector_data():
            return akshare.get_industry_sector_data()

        with ThreadPoolExecutor(max_workers=3) as executor:
            # 提交任务到线程池
            limit_up_data = executor.submit(get_limit_up_data)
            limit_down_data = executor.submit(get_limit_down_data)
            industry_sector_data = executor.submit(get_industry_sector_data)

            # 获取结果
            limit_up = limit_up_data.result()
            limit_down = limit_down_data.result()
            industry_sector = industry_sector_data.result()

        # 统计涨停和跌停板的行业分布
        limit_up_stat = limit_up['所属行业'].value_counts().reset_index()
        limit_up_stat.columns = ['行业', '涨停数']

        limit_down_stat = limit_down['所属行业'].value_counts().reset_index()
        limit_down_stat.columns = ['行业', '跌停数']

        # 合并统计数据
        statistics = pd.merge(limit_up_stat, limit_down_stat, on='行业', how='outer').fillna(0)
        statistics['涨停数'] = statistics['涨停数'].astype(int)
        statistics['跌停数'] = statistics['跌停数'].astype(int)


        # 获取板块信息
        industry_sector.rename(columns={'板块名称': '行业'}, inplace=True)  # 统一列名
        industry_sector = industry_sector[['板块代码', '行业', '上涨家数', '下跌家数']]  # 选择需要的字段
        industry_sector = industry_sector.copy()  # 确保操作在副本上进行
        industry_sector['股票总数'] = industry_sector['上涨家数'] + industry_sector['下跌家数']

        # 将板块信息合并到统计汇总中
        statistics = pd.merge(industry_sector, statistics, on='行业', how='left').fillna(0)
        statistics['涨停数'] = statistics['涨停数'].astype(int)
        statistics['跌停数'] = statistics['跌停数'].astype(int)

        # 涨停百分比
        # statistics['涨停百分比'] = np.where(statistics['股票总数'] == 0, 0, (statistics['涨停数'] / statistics['股票总数']))
        # statistics['跌停百分比'] = np.where(statistics['股票总数'] == 0, 0, (statistics['跌停数'] / statistics['股票总数']))

        # 定义期望的列顺序
        cols_order = ['板块代码', '行业', '股票总数', '涨停数', '跌停数']
        statistics = statistics[cols_order]
        # 创建Excel文件
        file_name = "板块涨跌停数据分析.xlsx"
        wb = Workbook()

        # Sheet1: 写入统计总结
        sheet1 = wb.active
        sheet1.title = "统计总结"

        # 写入列名
        sheet1.append(cols_order)

        # 写入统计数据
        for _, row in statistics.iterrows():
            sheet1.append(row.tolist())

        # Sheet2: 写入limit_up
        sheet2 = wb.create_sheet(title="涨停板数据")
        sheet2.append(limit_up.columns.tolist())  # 写入列名
        for row in limit_up.itertuples(index=False):
            sheet2.append(row)

        # Sheet3: 写入跌停数据
        sheet3 = wb.create_sheet(title="跌停板数据")
        sheet3.append(limit_down.columns.tolist())  # 写入列名
        for row in limit_down.itertuples(index=False):
            sheet3.append(row)

        # 保存Excel文件
        wb.save(file_name)

        print(f"板块数据结合涨跌停数据及统计结果已成功写入到 {file_name}！")
        """